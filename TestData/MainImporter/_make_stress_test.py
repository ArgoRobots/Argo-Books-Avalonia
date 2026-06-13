"""
Generates a deliberately complex, realistic-but-messy workbook for stress-testing the
experimental AI layout interpretation. Each sheet targets a different hard layout so a
single live import gives a clear read on capabilities.

Run:  python TestData/MainImporter/_make_stress_test.py
Out:  TestData/MainImporter/complex_stress_test.xlsx
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

OUT = os.path.join(os.path.dirname(__file__), "complex_stress_test.xlsx")

wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Revenue as a month cross-tab, buried under a title + note + blanks,
# with a trailing "Total" column AND a trailing total row (both should be ignored).
# Correct interpretation: pivot to one dated Revenue row per (product, month) = 12 rows.
# ─────────────────────────────────────────────────────────────────────────────
s1 = wb.active
s1.title = "Sales"
s1["A1"] = "Acme Trading Co.  -  Q1 2024 Sales by Product"
s1["A1"].font = Font(bold=True, size=14)
s1.merge_cells("A1:E1")
s1["A3"] = "All figures in USD. Prepared by J. Smith, 2024-04-02."
# row 4 blank
s1.append([])  # keeps row cursor sane; explicit cells below
hdr = 5
s1.cell(hdr, 1, "Product")
s1.cell(hdr, 2, "Jan")
s1.cell(hdr, 3, "Feb")
s1.cell(hdr, 4, "Mar")
s1.cell(hdr, 5, "Total")
for c in range(1, 6):
    s1.cell(hdr, c).font = Font(bold=True)

rows = [
    ("Widget",  1000, 1200, 1100),
    ("Gadget",  3000, 2500, 2800),
    ("Gizmo",    500,  750,  600),
    ("Sprocket", 200,  180,  220),
]
r = hdr + 1
for name, jan, feb, mar in rows:
    s1.cell(r, 1, name)
    s1.cell(r, 2, jan)
    s1.cell(r, 3, feb)
    s1.cell(r, 4, mar)
    s1.cell(r, 5, jan + feb + mar)  # trailing Total column (should NOT become revenue)
    r += 1
# blank separator row, then a grand-total row (should NOT be imported)
r += 1
s1.cell(r, 1, "TOTAL")
s1.cell(r, 2, sum(x[1] for x in rows))
s1.cell(r, 3, sum(x[2] for x in rows))
s1.cell(r, 4, sum(x[3] for x in rows))
s1.cell(r, 5, sum(x[1] + x[2] + x[3] for x in rows))

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Expenses report. Clean-ish long-form table but wrapped in a merged title,
# a blank row, a mid-table "Subtotal" ragged row, a blank separator, and a
# "Grand Total" row. Correct interpretation: 6 Expense rows, totals ignored.
# ─────────────────────────────────────────────────────────────────────────────
s2 = wb.create_sheet("Expenses")
s2["A1"] = "Operating Expenses - First Quarter"
s2["A1"].font = Font(bold=True, size=13)
s2.merge_cells("A1:D1")
# row 2 blank
s2.cell(3, 1, "Date");  s2.cell(3, 2, "Vendor")
s2.cell(3, 3, "Category"); s2.cell(3, 4, "Amount")
for c in range(1, 5):
    s2.cell(3, c).font = Font(bold=True)

exp = [
    (datetime(2024, 1, 5),  "Staples",       "Office Supplies", 142.50),
    (datetime(2024, 1, 18), "Verizon",       "Utilities",       210.00),
    (datetime(2024, 2, 2),  "City Power",     "Utilities",       340.75),
]
r = 4
for d, v, cat, amt in exp:
    s2.cell(r, 1, d); s2.cell(r, 1).number_format = "yyyy-mm-dd"
    s2.cell(r, 2, v); s2.cell(r, 3, cat); s2.cell(r, 4, amt)
    r += 1
# ragged subtotal row mid-table
s2.cell(r, 1, "Subtotal"); s2.cell(r, 4, sum(e[3] for e in exp)); r += 1
exp2 = [
    (datetime(2024, 2, 20), "Amazon",        "Office Supplies", 88.20),
    (datetime(2024, 3, 4),  "WeWork",        "Rent",           1500.00),
    (datetime(2024, 3, 15), "Adobe",         "Software",         59.99),
]
for d, v, cat, amt in exp2:
    s2.cell(r, 1, d); s2.cell(r, 1).number_format = "yyyy-mm-dd"
    s2.cell(r, 2, v); s2.cell(r, 3, cat); s2.cell(r, 4, amt)
    r += 1
r += 1  # blank
s2.cell(r, 1, "Grand Total")
s2.cell(r, 4, sum(e[3] for e in exp) + sum(e[3] for e in exp2))

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Customers stored TRANSPOSED - fields run down column A, each customer is
# its own COLUMN. The hardest layout here. Correct interpretation: transpose so each
# column becomes a Customer record (3 customers).
# ─────────────────────────────────────────────────────────────────────────────
s3 = wb.create_sheet("Clients")
s3.cell(1, 1, "Field").font = Font(bold=True)
s3.cell(1, 2, "Customer A")
s3.cell(1, 3, "Customer B")
s3.cell(1, 4, "Customer C")
fields = [
    ("Name",    "Acme Corp",        "Globex LLC",        "Initech Inc"),
    ("Email",   "ap@acme.com",      "billing@globex.com", "pay@initech.com"),
    ("Phone",   "555-0101",         "555-0202",          "555-0303"),
    ("City",    "Springfield",      "Ogdenville",        "North Haverbrook"),
]
r = 2
for label, a, b, c in fields:
    s3.cell(r, 1, label).font = Font(bold=True)
    s3.cell(r, 2, a); s3.cell(r, 3, b); s3.cell(r, 4, c)
    r += 1

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: A genuinely unsupported sheet (project notes) - should be reported as
# "cannot import", confirming the dialog still behaves correctly end-to-end.
# ─────────────────────────────────────────────────────────────────────────────
s4 = wb.create_sheet("Notes")
s4.cell(1, 1, "Project")
s4.cell(1, 2, "Owner")
s4.cell(1, 3, "Status")
s4.cell(2, 1, "Website redesign"); s4.cell(2, 2, "Dana"); s4.cell(2, 3, "In progress")
s4.cell(3, 1, "Q2 budget");        s4.cell(3, 2, "Lee");  s4.cell(3, 3, "Planned")

wb.save(OUT)
print("wrote", OUT)
