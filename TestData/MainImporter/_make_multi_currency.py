"""
Generates a multi-currency sample workbook for testing in-cell currency detection.

The amount cells are REAL NUMBERS carrying their currency via the cell's number FORMAT
(how a real business stores them), so the analyzer sees a clean numeric column and routes
the sheet to Tier 1 (direct mapping); the currency is read from the displayed format by the
deterministic pre-pass. This mirrors realistic spreadsheets far better than text amounts.

Currencies exercised:
  - £ / €     unambiguous symbol  -> GBP / EUR
  - "CAD"     explicit code in format suffix -> CAD
  - $ and ¥   ambiguous symbol    -> ONE prompt each
  - plain     no currency         -> company currency

Run:  python TestData/MainImporter/_make_multi_currency.py
Out:  TestData/MainImporter/multi_currency_sample.xlsx
"""
import os
from openpyxl import Workbook

OUT = os.path.join(os.path.dirname(__file__), "multi_currency_sample.xlsx")

GBP = '"£"#,##0.00'         # £1,200.00
EUR = '"€"#,##0.00'         # €800.00
CAD = '#,##0.00" CAD"'           # 1,500.00 CAD   (explicit code -> no prompt)
USD = '"$"#,##0.00'              # $950.00        (ambiguous $)
JPY = '"¥"#,##0'            # ¥95000         (ambiguous ¥)
PLAIN = '#,##0.00'               # 500.00         (no currency)

wb = Workbook()

# Each row is a single line at qty 1, so Unit Price == Total. Both columns carry the amount (and
# its currency format) so the imported record is internally consistent: the line item matches the
# stored total and the edit modal does not warn.

# ── Sheet 1: Revenue ──────────────────────────────────────────────────────────
s = wb.active
s.title = "Sales"
s.append(["ID", "Date", "Customer", "Product", "Unit Price", "Total"])
rows = [
    ("R1", "2026-01-05", "Acme UK",      "Widget",   1200.00, GBP),   # GBP
    ("R2", "2026-01-09", "Globex EU",    "Gadget",    800.00, EUR),   # EUR
    ("R3", "2026-01-14", "Maple Co",     "Gizmo",    1500.00, CAD),   # CAD (explicit)
    ("R4", "2026-01-18", "Yen Traders",  "Sprocket", 95000,   JPY),   # ambiguous ¥
    ("R5", "2026-01-22", "Dollar Store", "Widget",    950.00, USD),   # ambiguous $
    ("R6", "2026-01-27", "Local LLC",    "Gadget",    500.00, PLAIN), # company currency
]
for r in rows:
    s.append([r[0], r[1], r[2], r[3], r[4], r[4]])  # Unit Price + Total
    s.cell(row=s.max_row, column=5).number_format = r[5]
    s.cell(row=s.max_row, column=6).number_format = r[5]

# ── Sheet 2: Expenses ─────────────────────────────────────────────────────────
e = wb.create_sheet("Expenses")
e.append(["ID", "Date", "Supplier ID", "Description", "Unit Price", "Total"])
erows = [
    ("E1", "2026-01-06", "SUP-1", "London office supplies", 140.00, GBP),   # GBP
    ("E2", "2026-01-12", "SUP-2", "Toronto utilities",      320.00, CAD),   # CAD
    ("E3", "2026-01-20", "SUP-3", "Misc",                    88.00, USD),   # ambiguous $
    ("E4", "2026-01-28", "SUP-4", "Domestic rent",         1500.00, PLAIN), # company currency
]
for r in erows:
    e.append([r[0], r[1], r[2], r[3], r[4], r[4]])
    e.cell(row=e.max_row, column=5).number_format = r[5]
    e.cell(row=e.max_row, column=6).number_format = r[5]

wb.save(OUT)
print("wrote", OUT)
